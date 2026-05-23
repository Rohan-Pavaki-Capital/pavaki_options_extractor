"""
JSON to Excel Converter for Share-Based Compensation Data
==========================================================

CLEAN VERSION: Only shows fields/columns/rows that have actual data.
No null placeholders, no "—" for missing fields.

Takes the structured JSON output from the options extractor and generates
a professional, analyst-ready Excel workbook.

Usage:
    python json_to_excel.py <input.json> [output.xlsx]

    # Or as a library:
    from json_to_excel import build_workbook
    build_workbook("extraction.json", "report.xlsx")

Requirements:
    pip install openpyxl
"""

import json
import sys
import argparse
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: pip install openpyxl")


# ═════════════════════════════════════════════════════════════════════════════
# STYLES
# ═════════════════════════════════════════════════════════════════════════════

FONT_NAME = "Arial"

TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", start_color="1F4E78")

H1_FONT = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
H1_FILL = PatternFill("solid", start_color="2E75B6")

H2_FONT = Font(name=FONT_NAME, size=11, bold=True, color="000000")
H2_FILL = PatternFill("solid", start_color="DDEBF7")

LABEL_FONT = Font(name=FONT_NAME, size=10, bold=True)
DATA_FONT = Font(name=FONT_NAME, size=10)
ITALIC_FONT = Font(name=FONT_NAME, size=9, italic=True, color="595959")
TOTAL_FONT = Font(name=FONT_NAME, size=10, bold=True)

TOTAL_FILL = PatternFill("solid", start_color="FFF2CC")
PRIOR_FILL = PatternFill("solid", start_color="F2F2F2")
WARN_FILL = PatternFill("solid", start_color="FFEB9C")
GOOD_FILL = PatternFill("solid", start_color="C6EFCE")
BAD_FILL = PatternFill("solid", start_color="FFC7CE")

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

PLAN_COLORS = ["FFF2CC", "DAEEF3", "E2EFDA", "FCE4D6", "EDEDED", "DEEBF6", "FFE699", "C5E0B4"]


# ═════════════════════════════════════════════════════════════════════════════
# STYLING HELPERS
# ═════════════════════════════════════════════════════════════════════════════

def style_title(cell):
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = CENTER


def style_h1(cell):
    cell.font = H1_FONT
    cell.fill = H1_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def style_h2(cell):
    cell.font = H2_FONT
    cell.fill = H2_FILL
    cell.alignment = LEFT


def style_label(cell):
    cell.font = LABEL_FONT
    cell.alignment = LEFT
    cell.border = THIN_BORDER


def style_data(cell, fmt=None):
    cell.font = DATA_FONT
    cell.alignment = RIGHT
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt


def style_total(cell, fmt=None):
    cell.font = TOTAL_FONT
    cell.alignment = RIGHT
    cell.border = THIN_BORDER
    cell.fill = TOTAL_FILL
    if fmt:
        cell.number_format = fmt


def set_widths(sheet, widths):
    for i, w in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(i)].width = w


def safe_num(val, default=0):
    if val is None:
        return default
    try:
        return float(val) if isinstance(val, str) else val
    except (ValueError, TypeError):
        return default


def has_value(val):
    """Return True if a field has a meaningful value (not None, not empty string)."""
    if val is None:
        return False
    if isinstance(val, str) and val.strip() == "":
        return False
    return True


def get_plan_color(idx):
    return PLAN_COLORS[idx % len(PLAN_COLORS)]


def has_rollforward_data(plan):
    fields = ["opening_balance", "granted", "exercised", "forfeited_or_lapsed", "closing_balance"]
    return any(plan.get(f) is not None for f in fields)


def has_any_field(plan, fields):
    """Check if plan has ANY of the listed fields populated."""
    return any(plan.get(f) is not None for f in fields)


def get_currency_symbol(currency):
    symbols = {
        "GBP": "£", "USD": "$", "EUR": "€", "JPY": "¥",
        "SGD": "S$", "HKD": "HK$", "AUD": "A$", "CAD": "C$",
        "CNY": "¥", "KRW": "₩", "INR": "₹", "CHF": "CHF",
    }
    return symbols.get(currency, currency or "")


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 1: EXECUTIVE SUMMARY
# ═════════════════════════════════════════════════════════════════════════════

def build_executive_summary(wb, data):
    s = wb.create_sheet("1. Executive Summary")
    plans = data.get("plans", [])
    company = data.get("company_name") or "Share-Based Compensation Report"

    # Title
    s.merge_cells("A1:G1")
    s["A1"] = f"{company.upper()} — SHARE-BASED COMPENSATION ANALYSIS"
    style_title(s["A1"])
    s.row_dimensions[1].height = 30

    # Build info list — only include fields with data
    row = 3
    s[f"A{row}"] = "Report Information"
    style_h2(s[f"A{row}"])
    s.merge_cells(f"A{row}:G{row}")

    info = []
    if has_value(data.get("company_name")):
        info.append(("Company", data["company_name"]))
    if has_value(data.get("report_period")):
        info.append(("Report Period", data["report_period"]))
    if has_value(data.get("currency")):
        currency = data["currency"]
        info.append(("Currency", f"{currency} ({get_currency_symbol(currency)})"))
    if has_value(data.get("reporting_standard")):
        info.append(("Reporting Standard", data["reporting_standard"]))

    info.append(("Total Plans Disclosed", len(plans)))
    rf_count = sum(1 for p in plans if has_rollforward_data(p))
    if rf_count > 0:
        info.append(("Plans with Roll-forward Data", rf_count))

    info.append(("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")))

    if "_meta" in data:
        meta = data["_meta"]
        if has_value(meta.get("source_pdf")):
            info.append(("Source PDF", meta["source_pdf"]))
        if meta.get("cost", {}).get("total_cost_usd") is not None:
            info.append(("Extraction Cost", f"${meta['cost']['total_cost_usd']:.4f}"))

    for k, v in info:
        row += 1
        s[f"A{row}"] = k
        s[f"A{row}"].font = LABEL_FONT
        s[f"B{row}"] = v
        s[f"B{row}"].font = DATA_FONT
        s.merge_cells(f"B{row}:G{row}")

    # Plan summary table
    summary_plans = [p for p in plans if p.get("closing_balance") is not None or p.get("total_contingent_awards") is not None]

    if summary_plans:
        row += 2
        s[f"A{row}"] = "Plan Summary at a Glance"
        style_h2(s[f"A{row}"])
        s.merge_cells(f"A{row}:G{row}")

        # Determine which columns have ANY data
        has_ex_price_col = any(p.get("weighted_avg_exercise_price") is not None for p in summary_plans)
        has_fv_col = any(p.get("valuation_inputs", {}) and p.get("valuation_inputs", {}).get("fair_value_per_option") is not None for p in summary_plans)
        has_vesting_col = any(p.get("vesting_period_years") is not None for p in summary_plans)

        # Build dynamic headers
        row += 1
        headers = ["Plan", "Type", "Closing Balance", "% of Total"]
        if has_ex_price_col:
            headers.append("Avg Exercise Price")
        if has_fv_col:
            headers.append("Fair Value/Option")
        if has_vesting_col:
            headers.append("Vesting (yrs)")

        for i, h in enumerate(headers, 1):
            style_h1(s.cell(row=row, column=i, value=h))
        s.row_dimensions[row].height = 30

        start = row + 1
        for i, plan in enumerate(summary_plans):
            r = start + i
            closing = plan.get("closing_balance") or plan.get("total_contingent_awards") or 0

            col = 1
            s.cell(row=r, column=col, value=plan.get("plan_name", "—"))
            style_label(s.cell(row=r, column=col))
            col += 1

            s.cell(row=r, column=col, value=plan.get("plan_type", "—"))
            style_data(s.cell(row=r, column=col))
            s.cell(row=r, column=col).alignment = LEFT
            col += 1

            s.cell(row=r, column=col, value=closing)
            style_data(s.cell(row=r, column=col), "#,##0")
            closing_col_letter = get_column_letter(col)
            col += 1

            s.cell(row=r, column=col, value=f"={closing_col_letter}{r}/SUM(${closing_col_letter}${start}:${closing_col_letter}${start + len(summary_plans) - 1})")
            style_data(s.cell(row=r, column=col), "0.0%")
            col += 1

            if has_ex_price_col:
                ex_price = plan.get("weighted_avg_exercise_price")
                if ex_price is not None:
                    s.cell(row=r, column=col, value=ex_price)
                    style_data(s.cell(row=r, column=col), "#,##0.00")
                else:
                    style_data(s.cell(row=r, column=col))
                col += 1

            if has_fv_col:
                fv = None
                if plan.get("valuation_inputs"):
                    fv = plan["valuation_inputs"].get("fair_value_per_option")
                if fv is not None:
                    s.cell(row=r, column=col, value=fv)
                    style_data(s.cell(row=r, column=col), "#,##0.00")
                else:
                    style_data(s.cell(row=r, column=col))
                col += 1

            if has_vesting_col:
                vesting = plan.get("vesting_period_years")
                if vesting is not None:
                    s.cell(row=r, column=col, value=vesting)
                    style_data(s.cell(row=r, column=col), "0")
                else:
                    style_data(s.cell(row=r, column=col))
                col += 1

        # Total row
        total_row = start + len(summary_plans)
        s.cell(row=total_row, column=1, value="TOTAL")
        style_total(s.cell(row=total_row, column=1))
        s.cell(row=total_row, column=1).alignment = LEFT
        style_total(s.cell(row=total_row, column=2))
        s.cell(row=total_row, column=3, value=f"=SUM(C{start}:C{start + len(summary_plans) - 1})")
        style_total(s.cell(row=total_row, column=3), "#,##0")
        s.cell(row=total_row, column=4, value="100.0%")
        style_total(s.cell(row=total_row, column=4), "0.0%")
        for c in range(5, len(headers) + 1):
            style_total(s.cell(row=total_row, column=c))
        row = total_row

    # Key Insights
    row += 3
    s[f"A{row}"] = "Key Insights"
    style_h2(s[f"A{row}"])
    s.merge_cells(f"A{row}:G{row}")

    total_closing = sum(safe_num(p.get("closing_balance")) or safe_num(p.get("total_contingent_awards")) for p in plans)
    nil_cost_count = sum(1 for p in plans if p.get("is_nil_cost") is True)
    cash_settled_count = sum(1 for p in plans if p.get("is_cash_settled") is True)

    insights = []
    if total_closing > 0 and summary_plans:
        insights.append(("• Total awards outstanding:", f"{total_closing:,.0f} shares across {len(summary_plans)} plan(s)"))

        largest_plan = max(summary_plans, key=lambda p: safe_num(p.get("closing_balance")) or safe_num(p.get("total_contingent_awards")))
        largest_closing = safe_num(largest_plan.get("closing_balance")) or safe_num(largest_plan.get("total_contingent_awards"))
        pct = (largest_closing / total_closing * 100) if total_closing > 0 else 0
        if len(summary_plans) > 1:
            insights.append(("• Largest plan:", f"{largest_plan.get('plan_name', '?')} — {pct:.1f}% of total"))

    if nil_cost_count > 0:
        insights.append(("• Nil-cost plans:", f"{nil_cost_count} of {len(plans)}"))
    if cash_settled_count > 0:
        insights.append(("• Cash-settled plans:", f"{cash_settled_count} of {len(plans)}"))

    models = set()
    for p in plans:
        if has_value(p.get("valuation_model")):
            models.add(p["valuation_model"])
    if models:
        insights.append(("• Valuation model(s):", ", ".join(sorted(models))))

    if insights:
        for k, v in insights:
            row += 1
            s[f"A{row}"] = k
            s[f"A{row}"].font = LABEL_FONT
            s.merge_cells(f"A{row}:B{row}")
            s[f"C{row}"] = v
            s[f"C{row}"].font = DATA_FONT
            s.merge_cells(f"C{row}:G{row}")
    else:
        row += 1
        s[f"A{row}"] = "No quantitative insights available from extracted data."
        s[f"A{row}"].font = ITALIC_FONT
        s.merge_cells(f"A{row}:G{row}")

    set_widths(s, [22, 30, 18, 14, 18, 16, 14])


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 2: PLAN ROLL-FORWARD
# ═════════════════════════════════════════════════════════════════════════════

def build_rollforward(wb, data):
    s = wb.create_sheet("2. Plan Roll-Forward")
    plans = data.get("plans", [])

    # Title
    s.merge_cells("A1:K1")
    s["A1"] = f"PLAN ROLL-FORWARD — {data.get('report_period', 'Current Year')}"
    style_title(s["A1"])
    s.row_dimensions[1].height = 28

    rf_plans = [p for p in plans if has_rollforward_data(p)]

    if not rf_plans:
        s["A3"] = "No plans with roll-forward data found."
        s.merge_cells("A3:K3")
        s["A3"].font = ITALIC_FONT
        set_widths(s, [22, 24, 14, 12, 12, 16, 14, 12, 18, 16, 14])
        return

    # Determine which columns have ANY data across all plans
    has_granted = any(p.get("granted") is not None for p in rf_plans)
    has_exercised = any(p.get("exercised") is not None for p in rf_plans)
    has_lapsed = any(p.get("forfeited_or_lapsed") is not None for p in rf_plans)
    has_vested = any(p.get("vested") is not None for p in rf_plans)
    has_ex_price = any(p.get("weighted_avg_exercise_price") is not None for p in rf_plans)
    has_fv = any(p.get("valuation_inputs", {}) and p.get("valuation_inputs", {}).get("fair_value_per_option") is not None for p in rf_plans)

    # Build dynamic column structure
    columns = [("Plan", 1), ("Plan Type", 1), ("Opening", 1)]
    if has_granted:
        columns.append(("Granted", 1))
    if has_exercised:
        columns.append(("Exercised", 1))
    if has_lapsed:
        columns.append(("Lapsed/Forfeited", 1))
    if has_vested:
        columns.append(("Vested", 1))
    columns.append(("Closing", 1))
    columns.append(("Net Δ", 1))
    if has_ex_price:
        columns.append(("Wtd Avg Exercise Price", 1))
    if has_fv:
        columns.append(("Fair Value/Option", 1))
    columns.append(("Math Check", 1))

    # Map field name → column index for easy lookup
    col_map = {}
    for i, (name, _) in enumerate(columns, 1):
        col_map[name] = i

    # Current Year section
    s["A3"] = "Current Year Movement"
    style_h2(s["A3"])
    s.merge_cells(f"A3:{get_column_letter(len(columns))}3")

    for i, (h, _) in enumerate(columns, 1):
        style_h1(s.cell(row=4, column=i, value=h))
    s.row_dimensions[4].height = 35

    # Render plans
    for i, plan in enumerate(rf_plans):
        r = 5 + i

        opening_col = col_map["Opening"]
        closing_col = col_map["Closing"]
        net_col = col_map["Net Δ"]
        math_col = col_map["Math Check"]

        s.cell(row=r, column=col_map["Plan"], value=plan.get("plan_name", "—"))
        style_label(s.cell(row=r, column=col_map["Plan"]))

        s.cell(row=r, column=col_map["Plan Type"], value=plan.get("plan_type", "—"))
        style_data(s.cell(row=r, column=col_map["Plan Type"]))
        s.cell(row=r, column=col_map["Plan Type"]).alignment = LEFT

        # Opening
        opening = plan.get("opening_balance")
        if opening is not None:
            s.cell(row=r, column=opening_col, value=opening)
            style_data(s.cell(row=r, column=opening_col), "#,##0;(#,##0);-")
        else:
            style_data(s.cell(row=r, column=opening_col))

        # Conditional columns
        if has_granted:
            g = plan.get("granted")
            if g is not None:
                s.cell(row=r, column=col_map["Granted"], value=g)
                style_data(s.cell(row=r, column=col_map["Granted"]), "#,##0;(#,##0);-")
            else:
                style_data(s.cell(row=r, column=col_map["Granted"]))

        if has_exercised:
            e = plan.get("exercised")
            if e is not None:
                s.cell(row=r, column=col_map["Exercised"], value=e)
                style_data(s.cell(row=r, column=col_map["Exercised"]), "#,##0;(#,##0);-")
            else:
                style_data(s.cell(row=r, column=col_map["Exercised"]))

        if has_lapsed:
            l = plan.get("forfeited_or_lapsed")
            if l is not None:
                s.cell(row=r, column=col_map["Lapsed/Forfeited"], value=l)
                style_data(s.cell(row=r, column=col_map["Lapsed/Forfeited"]), "#,##0;(#,##0);-")
            else:
                style_data(s.cell(row=r, column=col_map["Lapsed/Forfeited"]))

        if has_vested:
            v = plan.get("vested")
            if v is not None:
                s.cell(row=r, column=col_map["Vested"], value=v)
                style_data(s.cell(row=r, column=col_map["Vested"]), "#,##0;(#,##0);-")
            else:
                style_data(s.cell(row=r, column=col_map["Vested"]))

        # Closing
        closing = plan.get("closing_balance")
        if closing is not None:
            s.cell(row=r, column=closing_col, value=closing)
            style_data(s.cell(row=r, column=closing_col), "#,##0;(#,##0);-")
        else:
            style_data(s.cell(row=r, column=closing_col))

        # Net Δ formula
        opening_letter = get_column_letter(opening_col)
        closing_letter = get_column_letter(closing_col)
        s.cell(row=r, column=net_col, value=f"=IFERROR({closing_letter}{r}-{opening_letter}{r},\"\")")
        style_data(s.cell(row=r, column=net_col), "#,##0;(#,##0);-")

        if has_ex_price:
            ex = plan.get("weighted_avg_exercise_price")
            if ex is not None:
                s.cell(row=r, column=col_map["Wtd Avg Exercise Price"], value=ex)
                style_data(s.cell(row=r, column=col_map["Wtd Avg Exercise Price"]), "#,##0.00")
            else:
                style_data(s.cell(row=r, column=col_map["Wtd Avg Exercise Price"]))

        if has_fv:
            fv = None
            if plan.get("valuation_inputs"):
                fv = plan["valuation_inputs"].get("fair_value_per_option")
            if fv is not None:
                s.cell(row=r, column=col_map["Fair Value/Option"], value=fv)
                style_data(s.cell(row=r, column=col_map["Fair Value/Option"]), "#,##0.00")
            else:
                style_data(s.cell(row=r, column=col_map["Fair Value/Option"]))

        # Math check
        granted_letter = get_column_letter(col_map.get("Granted", 0)) if has_granted else None
        exercised_letter = get_column_letter(col_map.get("Exercised", 0)) if has_exercised else None
        lapsed_letter = get_column_letter(col_map.get("Lapsed/Forfeited", 0)) if has_lapsed else None

        # Build dynamic math formula
        formula_parts = [f"{opening_letter}{r}"]
        if granted_letter:
            formula_parts.append(f"+IFERROR({granted_letter}{r},0)")
        if exercised_letter:
            formula_parts.append(f"-IFERROR({exercised_letter}{r},0)")
        if lapsed_letter:
            formula_parts.append(f"-IFERROR({lapsed_letter}{r},0)")
        formula = "".join(formula_parts)

        s.cell(row=r, column=math_col,
               value=f'=IF(OR({opening_letter}{r}="",{closing_letter}{r}=""),"n/a",IF(ABS(({formula})-{closing_letter}{r})<=1,"✓ OK","✗ "&TEXT(({formula})-{closing_letter}{r},"#,##0")))')
        style_data(s.cell(row=r, column=math_col))
        s.cell(row=r, column=math_col).alignment = CENTER

    # Total row CY
    tr = 5 + len(rf_plans)
    s.cell(row=tr, column=col_map["Plan"], value="TOTAL")
    style_total(s.cell(row=tr, column=col_map["Plan"]))
    s.cell(row=tr, column=col_map["Plan"]).alignment = LEFT
    style_total(s.cell(row=tr, column=col_map["Plan Type"]))

    # Sum numeric columns
    sum_columns = ["Opening", "Closing", "Net Δ"]
    if has_granted:
        sum_columns.append("Granted")
    if has_exercised:
        sum_columns.append("Exercised")
    if has_lapsed:
        sum_columns.append("Lapsed/Forfeited")
    if has_vested:
        sum_columns.append("Vested")

    for col_name in sum_columns:
        col = col_map[col_name]
        letter = get_column_letter(col)
        s.cell(row=tr, column=col, value=f"=SUM({letter}5:{letter}{tr - 1})")
        style_total(s.cell(row=tr, column=col), "#,##0;(#,##0);-")

    # Style other cells in total row
    for col_name in ["Wtd Avg Exercise Price", "Fair Value/Option", "Math Check"]:
        if col_name in col_map:
            style_total(s.cell(row=tr, column=col_map[col_name]))

    # Prior year section
    py_plans = [p for p in rf_plans if p.get("prior_year") and any(p["prior_year"].get(k) is not None for k in ["opening_balance", "closing_balance", "granted", "exercised", "forfeited_or_lapsed"])]

    last_total_row = tr

    if py_plans:
        py_start = tr + 3
        s.cell(row=py_start, column=1, value="Prior Year (Comparative)")
        style_h2(s.cell(row=py_start, column=1))
        s.merge_cells(f"A{py_start}:{get_column_letter(len(columns))}{py_start}")

        # Determine PY columns (might differ from CY)
        py_has_granted = any(p["prior_year"].get("granted") is not None for p in py_plans)
        py_has_exercised = any(p["prior_year"].get("exercised") is not None for p in py_plans)
        py_has_lapsed = any(p["prior_year"].get("forfeited_or_lapsed") is not None for p in py_plans)
        py_has_vested = any(p["prior_year"].get("vested") is not None for p in py_plans)
        py_has_ex_price = any(p["prior_year"].get("weighted_avg_exercise_price") is not None for p in py_plans)

        # Use same column layout as CY for visual consistency
        py_start += 1
        for i, (h, _) in enumerate(columns, 1):
            style_h1(s.cell(row=py_start, column=i, value=h))
        s.row_dimensions[py_start].height = 35

        for i, plan in enumerate(py_plans):
            r = py_start + 1 + i
            py = plan["prior_year"]

            s.cell(row=r, column=col_map["Plan"], value=plan.get("plan_name", "—"))
            style_label(s.cell(row=r, column=col_map["Plan"]))
            s.cell(row=r, column=col_map["Plan"]).fill = PRIOR_FILL

            s.cell(row=r, column=col_map["Plan Type"], value=plan.get("plan_type", "—"))
            style_data(s.cell(row=r, column=col_map["Plan Type"]))
            s.cell(row=r, column=col_map["Plan Type"]).alignment = LEFT
            s.cell(row=r, column=col_map["Plan Type"]).fill = PRIOR_FILL

            for field_name, col_name in [
                ("opening_balance", "Opening"),
                ("granted", "Granted"),
                ("exercised", "Exercised"),
                ("forfeited_or_lapsed", "Lapsed/Forfeited"),
                ("vested", "Vested"),
                ("closing_balance", "Closing"),
            ]:
                if col_name in col_map:
                    val = py.get(field_name)
                    if val is not None:
                        s.cell(row=r, column=col_map[col_name], value=val)
                        style_data(s.cell(row=r, column=col_map[col_name]), "#,##0;(#,##0);-")
                    else:
                        style_data(s.cell(row=r, column=col_map[col_name]))
                    s.cell(row=r, column=col_map[col_name]).fill = PRIOR_FILL

            # Net Δ
            opening_letter = get_column_letter(col_map["Opening"])
            closing_letter = get_column_letter(col_map["Closing"])
            s.cell(row=r, column=col_map["Net Δ"], value=f"=IFERROR({closing_letter}{r}-{opening_letter}{r},\"\")")
            style_data(s.cell(row=r, column=col_map["Net Δ"]), "#,##0;(#,##0);-")
            s.cell(row=r, column=col_map["Net Δ"]).fill = PRIOR_FILL

            # PY Exercise price
            if "Wtd Avg Exercise Price" in col_map:
                py_ex = py.get("weighted_avg_exercise_price")
                if py_ex is not None:
                    s.cell(row=r, column=col_map["Wtd Avg Exercise Price"], value=py_ex)
                    style_data(s.cell(row=r, column=col_map["Wtd Avg Exercise Price"]), "#,##0.00")
                else:
                    style_data(s.cell(row=r, column=col_map["Wtd Avg Exercise Price"]))
                s.cell(row=r, column=col_map["Wtd Avg Exercise Price"]).fill = PRIOR_FILL

            if "Fair Value/Option" in col_map:
                style_data(s.cell(row=r, column=col_map["Fair Value/Option"]))
                s.cell(row=r, column=col_map["Fair Value/Option"]).fill = PRIOR_FILL

            # Math check
            granted_letter = get_column_letter(col_map.get("Granted", 0)) if has_granted else None
            exercised_letter = get_column_letter(col_map.get("Exercised", 0)) if has_exercised else None
            lapsed_letter = get_column_letter(col_map.get("Lapsed/Forfeited", 0)) if has_lapsed else None

            formula_parts = [f"{opening_letter}{r}"]
            if granted_letter:
                formula_parts.append(f"+IFERROR({granted_letter}{r},0)")
            if exercised_letter:
                formula_parts.append(f"-IFERROR({exercised_letter}{r},0)")
            if lapsed_letter:
                formula_parts.append(f"-IFERROR({lapsed_letter}{r},0)")
            formula = "".join(formula_parts)

            s.cell(row=r, column=col_map["Math Check"],
                   value=f'=IF(OR({opening_letter}{r}="",{closing_letter}{r}=""),"n/a",IF(ABS(({formula})-{closing_letter}{r})<=1,"✓ OK","✗ "&TEXT(({formula})-{closing_letter}{r},"#,##0")))')
            style_data(s.cell(row=r, column=col_map["Math Check"]))
            s.cell(row=r, column=col_map["Math Check"]).alignment = CENTER
            s.cell(row=r, column=col_map["Math Check"]).fill = PRIOR_FILL

        # Total prior year
        py_tr = py_start + 1 + len(py_plans)
        s.cell(row=py_tr, column=col_map["Plan"], value="TOTAL")
        style_total(s.cell(row=py_tr, column=col_map["Plan"]))
        s.cell(row=py_tr, column=col_map["Plan"]).alignment = LEFT
        style_total(s.cell(row=py_tr, column=col_map["Plan Type"]))

        for col_name in sum_columns:
            col = col_map[col_name]
            letter = get_column_letter(col)
            s.cell(row=py_tr, column=col, value=f"=SUM({letter}{py_start + 1}:{letter}{py_tr - 1})")
            style_total(s.cell(row=py_tr, column=col), "#,##0;(#,##0);-")

        for col_name in ["Wtd Avg Exercise Price", "Fair Value/Option", "Math Check"]:
            if col_name in col_map:
                style_total(s.cell(row=py_tr, column=col_map[col_name]))

        last_total_row = py_tr

    # Notes
    last_row = last_total_row + 3
    s[f"A{last_row}"] = "Notes:"
    s[f"A{last_row}"].font = Font(name=FONT_NAME, size=9, bold=True, italic=True)
    s.merge_cells(f"A{last_row}:{get_column_letter(len(columns))}{last_row}")

    notes = [
        "• Math Check verifies: Opening + Granted - Exercised - Lapsed = Closing (±1 unit tolerance)",
        "• Negative numbers shown in parentheses",
        "• Empty cells indicate data was not disclosed in source",
    ]
    for note in notes:
        last_row += 1
        s[f"A{last_row}"] = note
        s[f"A{last_row}"].font = ITALIC_FONT
        s.merge_cells(f"A{last_row}:{get_column_letter(len(columns))}{last_row}")

    # Dynamic column widths
    widths = [22, 24]  # Plan, Plan Type
    for col_name, _ in columns[2:]:
        if "Price" in col_name or "Value" in col_name:
            widths.append(18)
        elif "Math" in col_name:
            widths.append(14)
        elif col_name == "Net Δ":
            widths.append(12)
        else:
            widths.append(14)
    set_widths(s, widths)


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 3: TRANCHE DETAILS
# ═════════════════════════════════════════════════════════════════════════════

def build_tranches(wb, data):
    s = wb.create_sheet("3. Tranche Details")
    plans = data.get("plans", [])

    s.merge_cells("A1:G1")
    s["A1"] = "TRANCHE-LEVEL DETAIL — INDIVIDUAL GRANTS"
    style_title(s["A1"])
    s.row_dimensions[1].height = 28

    all_tranches = []
    plan_color_map = {}
    for i, plan in enumerate(plans):
        plan_name = plan.get("plan_name", "—")
        plan_color_map[plan_name] = get_plan_color(i)
        for tranche in (plan.get("tranches") or []):
            all_tranches.append((plan_name, plan.get("plan_type", "—"), tranche))

    if not all_tranches:
        s["A3"] = "No tranche-level data disclosed in source."
        s.merge_cells("A3:G3")
        s["A3"].font = ITALIC_FONT
        set_widths(s, [16, 14, 18, 18, 16, 18, 24])
        return

    # Determine which columns have data
    has_grant_price = any(t.get("grant_price") is not None for _, _, t in all_tranches)
    has_exercise_price = any(t.get("exercise_price") is not None for _, _, t in all_tranches)
    has_vesting = any(t.get("vesting_period_years") is not None for _, _, t in all_tranches)
    has_fv = any(t.get("fair_value_per_option") is not None for _, _, t in all_tranches)
    has_grant_date = any(t.get("grant_date") is not None for _, _, t in all_tranches)
    has_shares = any((t.get("shares_at_period_end") or t.get("shares_granted")) is not None for _, _, t in all_tranches)

    s["A3"] = "All Outstanding Tranches by Plan and Grant Date"
    style_h2(s["A3"])
    s.merge_cells("A3:G3")

    # Build dynamic columns
    columns = ["Plan"]
    if has_grant_date:
        columns.append("Grant Date")
    if has_shares:
        columns.append("Shares at Period End")
    if has_grant_price:
        columns.append("Grant Price")
    if has_exercise_price:
        columns.append("Exercise Price")
    if has_vesting:
        columns.append("Vesting (yrs)")
    if has_fv:
        columns.append("Fair Value/Option")

    col_map = {name: i + 1 for i, name in enumerate(columns)}

    for i, h in enumerate(columns, 1):
        style_h1(s.cell(row=4, column=i, value=h))
    s.row_dimensions[4].height = 30

    start_row = 5
    for i, (plan_name, plan_type, t) in enumerate(all_tranches):
        r = start_row + i

        s.cell(row=r, column=col_map["Plan"], value=plan_name)
        style_label(s.cell(row=r, column=col_map["Plan"]))
        s.cell(row=r, column=col_map["Plan"]).fill = PatternFill("solid", start_color=plan_color_map[plan_name])

        if has_grant_date:
            gd = t.get("grant_date")
            if gd is not None:
                s.cell(row=r, column=col_map["Grant Date"], value=gd)
                style_data(s.cell(row=r, column=col_map["Grant Date"]))
                s.cell(row=r, column=col_map["Grant Date"]).alignment = CENTER
            else:
                style_data(s.cell(row=r, column=col_map["Grant Date"]))

        if has_shares:
            shares = t.get("shares_at_period_end") or t.get("shares_granted")
            if shares is not None:
                s.cell(row=r, column=col_map["Shares at Period End"], value=shares)
                style_data(s.cell(row=r, column=col_map["Shares at Period End"]), "#,##0")
            else:
                style_data(s.cell(row=r, column=col_map["Shares at Period End"]))

        if has_grant_price:
            gp = t.get("grant_price")
            if gp is not None:
                s.cell(row=r, column=col_map["Grant Price"], value=gp)
                style_data(s.cell(row=r, column=col_map["Grant Price"]), "#,##0.00")
            else:
                style_data(s.cell(row=r, column=col_map["Grant Price"]))

        if has_exercise_price:
            ep = t.get("exercise_price")
            if ep is not None:
                s.cell(row=r, column=col_map["Exercise Price"], value=ep)
                style_data(s.cell(row=r, column=col_map["Exercise Price"]), "#,##0.00")
            else:
                style_data(s.cell(row=r, column=col_map["Exercise Price"]))

        if has_vesting:
            vp = t.get("vesting_period_years")
            if vp is not None:
                s.cell(row=r, column=col_map["Vesting (yrs)"], value=vp)
                style_data(s.cell(row=r, column=col_map["Vesting (yrs)"]), "0")
            else:
                style_data(s.cell(row=r, column=col_map["Vesting (yrs)"]))

        if has_fv:
            fv = t.get("fair_value_per_option")
            if fv is not None:
                s.cell(row=r, column=col_map["Fair Value/Option"], value=fv)
                style_data(s.cell(row=r, column=col_map["Fair Value/Option"]), "#,##0.00")
            else:
                style_data(s.cell(row=r, column=col_map["Fair Value/Option"]))

    # Total row (only sum shares column)
    tr_row = start_row + len(all_tranches)
    s.cell(row=tr_row, column=col_map["Plan"], value="TOTAL")
    style_total(s.cell(row=tr_row, column=col_map["Plan"]))
    s.cell(row=tr_row, column=col_map["Plan"]).alignment = LEFT

    for col_name in columns[1:]:
        col = col_map[col_name]
        if col_name == "Shares at Period End":
            letter = get_column_letter(col)
            s.cell(row=tr_row, column=col, value=f"=SUM({letter}{start_row}:{letter}{tr_row - 1})")
            style_total(s.cell(row=tr_row, column=col), "#,##0")
        else:
            style_total(s.cell(row=tr_row, column=col))

    # Subtotals by plan
    sub_row = tr_row + 2
    s.cell(row=sub_row, column=1, value="Subtotals by Plan")
    style_h2(s.cell(row=sub_row, column=1))
    s.merge_cells(f"A{sub_row}:{get_column_letter(len(columns))}{sub_row}")

    sub_row += 1
    sub_headers = ["Plan", "# Tranches", "Total Shares"]
    if has_grant_price:
        sub_headers.extend(["Min Grant Price", "Max Grant Price", "Wtd Avg Grant Price"])

    for i, h in enumerate(sub_headers, 1):
        style_h1(s.cell(row=sub_row, column=i, value=h))

    # Find row ranges per plan
    plan_ranges = {}
    current_plan = None
    range_start = None
    for i, (plan_name, _, _) in enumerate(all_tranches):
        excel_row = start_row + i
        if plan_name != current_plan:
            if current_plan is not None:
                plan_ranges[current_plan] = (range_start, excel_row - 1)
            current_plan = plan_name
            range_start = excel_row
    if current_plan is not None:
        plan_ranges[current_plan] = (range_start, start_row + len(all_tranches) - 1)

    sub_row += 1
    for i, (plan_name, (s_row, e_row)) in enumerate(plan_ranges.items()):
        r = sub_row + i
        s.cell(row=r, column=1, value=plan_name)
        style_label(s.cell(row=r, column=1))
        s.cell(row=r, column=1).fill = PatternFill("solid", start_color=plan_color_map[plan_name])

        s.cell(row=r, column=2, value=e_row - s_row + 1)
        style_data(s.cell(row=r, column=2), "0")

        if has_shares:
            shares_letter = get_column_letter(col_map["Shares at Period End"])
            s.cell(row=r, column=3, value=f"=SUM({shares_letter}{s_row}:{shares_letter}{e_row})")
            style_data(s.cell(row=r, column=3), "#,##0")
        else:
            style_data(s.cell(row=r, column=3))

        if has_grant_price:
            gp_letter = get_column_letter(col_map["Grant Price"])
            s.cell(row=r, column=4, value=f"=IFERROR(MIN({gp_letter}{s_row}:{gp_letter}{e_row}),\"\")")
            style_data(s.cell(row=r, column=4), "#,##0.00")
            s.cell(row=r, column=5, value=f"=IFERROR(MAX({gp_letter}{s_row}:{gp_letter}{e_row}),\"\")")
            style_data(s.cell(row=r, column=5), "#,##0.00")
            if has_shares:
                shares_letter = get_column_letter(col_map["Shares at Period End"])
                s.cell(row=r, column=6, value=f"=IFERROR(SUMPRODUCT({shares_letter}{s_row}:{shares_letter}{e_row},{gp_letter}{s_row}:{gp_letter}{e_row})/SUM({shares_letter}{s_row}:{shares_letter}{e_row}),\"\")")
                style_data(s.cell(row=r, column=6), "#,##0.00")

    # Set widths
    widths = [22]
    for col_name in columns[1:]:
        if "Date" in col_name:
            widths.append(18)
        elif "Shares" in col_name:
            widths.append(20)
        elif "Price" in col_name or "Value" in col_name:
            widths.append(16)
        else:
            widths.append(14)
    set_widths(s, widths)


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 4: VALUATION INPUTS
# ═════════════════════════════════════════════════════════════════════════════

def build_valuation_inputs(wb, data):
    s = wb.create_sheet("4. Valuation Inputs")
    plans = data.get("plans", [])

    s.merge_cells("A1:F1")
    s["A1"] = "VALUATION MODEL INPUTS"
    style_title(s["A1"])
    s.row_dimensions[1].height = 28

    val_plans = [p for p in plans if p.get("valuation_inputs") and any(
        p["valuation_inputs"].get(k) is not None for k in
        ["volatility_pct", "risk_free_rate_pct", "fair_value_per_option", "dividend_yield_pct",
         "expected_life_years", "stock_price", "strike_price"]
    )]

    if not val_plans:
        s["A3"] = "No valuation model inputs disclosed in source."
        s.merge_cells("A3:F3")
        s["A3"].font = ITALIC_FONT
        set_widths(s, [28, 16, 16, 16, 16, 16])
        return

    s["A3"] = "Black-Scholes / Monte Carlo Parameters by Plan"
    style_h2(s["A3"])
    s.merge_cells(f"A3:{get_column_letter(len(val_plans) + 1)}3")

    # Build header
    headers = ["Parameter"] + [p.get("plan_name", f"Plan {i+1}")[:25] for i, p in enumerate(val_plans)]
    for i, h in enumerate(headers, 1):
        style_h1(s.cell(row=4, column=i, value=h))
    s.row_dimensions[4].height = 25

    # Parameters with field name, format, type
    all_params = [
        ("Valuation Model", "valuation_model", None, "model"),
        ("Expected Volatility (%)", "volatility_pct", "0.00%", "pct"),
        ("Risk-Free Rate (%)", "risk_free_rate_pct", "0.00%", "pct"),
        ("Dividend Yield (%)", "dividend_yield_pct", "0.00%", "pct"),
        ("Expected Life (years)", "expected_life_years", "0.00", "num"),
        ("Stock Price at Grant", "stock_price", "#,##0.00", "num"),
        ("Strike Price", "strike_price", "#,##0.00", "num"),
        ("Fair Value per Option", "fair_value_per_option", "#,##0.00", "num"),
    ]

    # Filter: only include rows where AT LEAST ONE plan has the value
    param_fields = []
    for label, field, fmt, ftype in all_params:
        has_any = False
        for plan in val_plans:
            if ftype == "model":
                if plan.get("valuation_model") is not None:
                    has_any = True
                    break
            else:
                vi = plan.get("valuation_inputs") or {}
                if vi.get(field) is not None:
                    has_any = True
                    break
        if has_any:
            param_fields.append((label, field, fmt, ftype))

    for i, (label, field, fmt, ftype) in enumerate(param_fields):
        r = 5 + i
        s.cell(row=r, column=1, value=label)
        style_label(s.cell(row=r, column=1))

        for j, plan in enumerate(val_plans):
            col = 2 + j
            vi = plan.get("valuation_inputs") or {}

            if ftype == "model":
                val = plan.get("valuation_model")
            else:
                val = vi.get(field)

            if val is not None:
                if ftype == "model":
                    s.cell(row=r, column=col, value=val)
                    style_data(s.cell(row=r, column=col))
                    s.cell(row=r, column=col).alignment = CENTER
                elif ftype == "pct":
                    s.cell(row=r, column=col, value=val / 100 if isinstance(val, (int, float)) else val)
                    style_data(s.cell(row=r, column=col), "0.00%")
                else:
                    s.cell(row=r, column=col, value=val)
                    style_data(s.cell(row=r, column=col), fmt)
            else:
                style_data(s.cell(row=r, column=col))

    widths = [28] + [18] * len(val_plans)
    set_widths(s, widths)


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 5: PLAN DESCRIPTIONS
# ═════════════════════════════════════════════════════════════════════════════

def build_plan_descriptions(wb, data):
    s = wb.create_sheet("5. Plan Descriptions")
    plans = data.get("plans", [])

    s.merge_cells("A1:C1")
    s["A1"] = "PLAN TYPES & DESCRIPTIONS"
    style_title(s["A1"])
    s.row_dimensions[1].height = 28

    # Filter: only plans with name or description
    desc_plans = [p for p in plans if has_value(p.get("plan_name")) or has_value(p.get("plan_description"))]

    if not desc_plans:
        s["A3"] = "No plan descriptions disclosed in source."
        s.merge_cells("A3:C3")
        s["A3"].font = ITALIC_FONT
        set_widths(s, [28, 32, 75])
        return

    for i, h in enumerate(["Plan", "Plan Type", "Description / Key Features"], 1):
        style_h1(s.cell(row=3, column=i, value=h))
    s.row_dimensions[3].height = 25

    for i, plan in enumerate(desc_plans):
        r = 4 + i

        s.cell(row=r, column=1, value=plan.get("plan_name", "—"))
        style_label(s.cell(row=r, column=1))
        s.cell(row=r, column=1).alignment = LEFT_TOP

        plan_type_full = plan.get("plan_type", "—")
        type_extras = []
        if plan.get("is_nil_cost"):
            type_extras.append("Nil-cost")
        if plan.get("is_cash_settled") is True:
            type_extras.append("Cash-settled")
        elif plan.get("is_cash_settled") is False:
            type_extras.append("Equity-settled")
        if type_extras:
            plan_type_full = f"{plan_type_full} ({', '.join(type_extras)})"

        s.cell(row=r, column=2, value=plan_type_full)
        style_data(s.cell(row=r, column=2))
        s.cell(row=r, column=2).alignment = LEFT_TOP

        # Build description from available fields
        desc_parts = []
        if has_value(plan.get("plan_description")):
            desc_parts.append(plan["plan_description"])
        if has_value(plan.get("vesting_description")):
            desc_parts.append(f"Vesting: {plan['vesting_description']}.")
        if plan.get("vesting_period_years") is not None and not has_value(plan.get("vesting_description")):
            desc_parts.append(f"Vesting period: {plan['vesting_period_years']} years.")
        if has_value(plan.get("performance_conditions")):
            desc_parts.append(f"Performance conditions: {plan['performance_conditions']}.")
        if plan.get("performance_period_years") is not None:
            desc_parts.append(f"Performance period: {plan['performance_period_years']} years.")
        if plan.get("holding_period_years") is not None:
            desc_parts.append(f"Holding period: {plan['holding_period_years']} years post-vesting.")
        if plan.get("maximum_payout_pct") is not None:
            desc_parts.append(f"Maximum payout: {plan['maximum_payout_pct']}% of baseline.")

        description = " ".join(desc_parts) if desc_parts else "No description disclosed."

        s.cell(row=r, column=3, value=description)
        style_data(s.cell(row=r, column=3))
        s.cell(row=r, column=3).alignment = LEFT_TOP

        text_len = len(description)
        s.row_dimensions[r].height = max(30, min(150, text_len / 8))

    set_widths(s, [28, 32, 75])


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 6: KPIs & RATIOS
# ═════════════════════════════════════════════════════════════════════════════

def build_kpis(wb, data):
    s = wb.create_sheet("6. KPIs & Ratios")
    plans = data.get("plans", [])

    s.merge_cells("A1:F1")
    s["A1"] = "KEY PERFORMANCE INDICATORS & RATIOS"
    style_title(s["A1"])
    s.row_dimensions[1].height = 28

    rf_plans = [p for p in plans if has_rollforward_data(p)]
    if not rf_plans:
        s["A3"] = "No roll-forward data available for KPI calculation."
        s.merge_cells("A3:F3")
        s["A3"].font = ITALIC_FONT
        set_widths(s, [28, 18, 18, 18, 14, 38])
        return

    # Determine which metrics have data
    def sum_field(plans, field):
        return sum(safe_num(p.get(field)) for p in plans if p.get(field) is not None)

    def sum_py_field(plans, field):
        return sum(
            safe_num(p["prior_year"].get(field))
            for p in plans
            if p.get("prior_year") and p["prior_year"].get(field) is not None
        )

    all_kpis = [
        ("Total Opening Balance", "opening_balance", "Across all plans"),
        ("Total Granted", "granted", "New equity awards"),
        ("Total Exercised", "exercised", "Awards converted to shares"),
        ("Total Lapsed/Forfeited", "forfeited_or_lapsed", "Awards expired or forfeited"),
        ("Total Closing Balance", "closing_balance", "Outstanding at period end"),
    ]

    # Filter KPIs: only include if CY value > 0
    kpi_data = []
    cy_values = {}
    py_values = {}
    for metric, field, notes in all_kpis:
        cy = sum_field(rf_plans, field)
        py = sum_py_field(rf_plans, field)
        if cy > 0 or py > 0:
            kpi_data.append((metric, field, notes, cy, py))
            cy_values[field] = cy
            py_values[field] = py

    if not kpi_data:
        s["A3"] = "No quantitative KPI data available."
        s.merge_cells("A3:F3")
        s["A3"].font = ITALIC_FONT
        set_widths(s, [28, 18, 18, 18, 14, 38])
        return

    # Aggregate metrics
    s["A3"] = "Aggregate Activity Metrics"
    style_h2(s["A3"])
    s.merge_cells("A3:F3")

    has_py_data = any(py > 0 for _, _, _, _, py in kpi_data)

    headers = ["Metric", "Current Year"]
    if has_py_data:
        headers.extend(["Prior Year", "Δ Change", "% Change"])
    headers.append("Notes")

    for i, h in enumerate(headers, 1):
        style_h1(s.cell(row=4, column=i, value=h))
    s.row_dimensions[4].height = 25

    # Track row positions for formula references
    metric_rows = {}

    for i, (metric, field, notes, cy, py) in enumerate(kpi_data):
        r = 5 + i
        metric_rows[field] = r

        s.cell(row=r, column=1, value=metric)
        style_label(s.cell(row=r, column=1))

        s.cell(row=r, column=2, value=cy)
        style_data(s.cell(row=r, column=2), "#,##0;(#,##0);-")

        col_idx = 3
        if has_py_data:
            s.cell(row=r, column=col_idx, value=py if py > 0 else None)
            style_data(s.cell(row=r, column=col_idx), "#,##0;(#,##0);-")
            col_idx += 1

            s.cell(row=r, column=col_idx, value=f"=IFERROR(B{r}-C{r},\"\")")
            style_data(s.cell(row=r, column=col_idx), "#,##0;(#,##0);-")
            col_idx += 1

            s.cell(row=r, column=col_idx, value=f"=IFERROR((B{r}-C{r})/C{r},\"\")")
            style_data(s.cell(row=r, column=col_idx), "0.0%;(0.0%);-")
            col_idx += 1

        s.cell(row=r, column=col_idx, value=notes)
        style_data(s.cell(row=r, column=col_idx))
        s.cell(row=r, column=col_idx).alignment = LEFT
        s.cell(row=r, column=col_idx).font = ITALIC_FONT

    # Activity Rates (only if we have opening_balance)
    if "opening_balance" in metric_rows and cy_values.get("opening_balance", 0) > 0:
        row = 5 + len(kpi_data) + 2
        s.cell(row=row, column=1, value="Activity Rates (% of Opening Balance)")
        style_h2(s.cell(row=row, column=1))
        s.merge_cells(f"A{row}:{get_column_letter(len(headers))}{row}")

        row += 1
        for i, h in enumerate(headers, 1):
            style_h1(s.cell(row=row, column=i, value=h))
        s.row_dimensions[row].height = 25

        rates_to_show = []
        opening_row = metric_rows["opening_balance"]

        if "granted" in metric_rows:
            rates_to_show.append(("Grant Rate", metric_rows["granted"], "New grants as % of opening"))
        if "exercised" in metric_rows:
            rates_to_show.append(("Exercise Rate", metric_rows["exercised"], "Conversion rate"))
        if "forfeited_or_lapsed" in metric_rows:
            rates_to_show.append(("Lapse Rate", metric_rows["forfeited_or_lapsed"], "Forfeit/expiry rate"))
        if "closing_balance" in metric_rows:
            rates_to_show.append(("Retention Rate", metric_rows["closing_balance"], "Awards remaining outstanding"))

        row += 1
        for i, (metric, src_row, notes) in enumerate(rates_to_show):
            r = row + i
            s.cell(row=r, column=1, value=metric)
            style_label(s.cell(row=r, column=1))

            s.cell(row=r, column=2, value=f"=IFERROR(B{src_row}/B{opening_row},\"\")")
            style_data(s.cell(row=r, column=2), "0.0%;(0.0%);-")

            col_idx = 3
            if has_py_data:
                s.cell(row=r, column=col_idx, value=f"=IFERROR(C{src_row}/C{opening_row},\"\")")
                style_data(s.cell(row=r, column=col_idx), "0.0%;(0.0%);-")
                col_idx += 1

                s.cell(row=r, column=col_idx, value=f"=IFERROR(B{r}-C{r},\"\")")
                style_data(s.cell(row=r, column=col_idx), "0.0%;(0.0%);-")
                col_idx += 1

                style_data(s.cell(row=r, column=col_idx))
                col_idx += 1

            s.cell(row=r, column=col_idx, value=notes)
            style_data(s.cell(row=r, column=col_idx))
            s.cell(row=r, column=col_idx).alignment = LEFT
            s.cell(row=r, column=col_idx).font = ITALIC_FONT

        last_row = row + len(rates_to_show)
    else:
        last_row = 5 + len(kpi_data)

    # Plan composition
    summary_plans = [p for p in plans if p.get("closing_balance") is not None or p.get("total_contingent_awards") is not None]

    if summary_plans:
        row = last_row + 3
        s.cell(row=row, column=1, value="Plan Composition (Current Year Closing)")
        style_h2(s.cell(row=row, column=1))
        s.merge_cells(f"A{row}:{get_column_letter(len(headers))}{row}")

        row += 1
        # Determine which extra columns to show
        has_settlement = any(p.get("is_cash_settled") is not None for p in summary_plans)

        comp_headers = ["Plan", "Closing Balance", "% of Total", "Plan Type"]
        if has_settlement:
            comp_headers.append("Settlement")

        for i, h in enumerate(comp_headers, 1):
            style_h1(s.cell(row=row, column=i, value=h))
        s.row_dimensions[row].height = 25

        start = row + 1
        for i, plan in enumerate(summary_plans):
            r = start + i
            closing = plan.get("closing_balance") or plan.get("total_contingent_awards") or 0

            s.cell(row=r, column=1, value=plan.get("plan_name", "—"))
            style_label(s.cell(row=r, column=1))

            s.cell(row=r, column=2, value=closing)
            style_data(s.cell(row=r, column=2), "#,##0")

            s.cell(row=r, column=3, value=f"=B{r}/SUM($B${start}:$B${start + len(summary_plans) - 1})")
            style_data(s.cell(row=r, column=3), "0.0%")

            s.cell(row=r, column=4, value=plan.get("plan_type", "—"))
            style_data(s.cell(row=r, column=4))
            s.cell(row=r, column=4).alignment = LEFT

            if has_settlement:
                if plan.get("is_cash_settled") is True:
                    settlement = "Cash-settled"
                elif plan.get("is_cash_settled") is False:
                    settlement = "Equity-settled"
                else:
                    settlement = ""
                s.cell(row=r, column=5, value=settlement)
                style_data(s.cell(row=r, column=5))
                s.cell(row=r, column=5).alignment = LEFT

        # Total
        tr = start + len(summary_plans)
        s.cell(row=tr, column=1, value="TOTAL")
        style_total(s.cell(row=tr, column=1))
        s.cell(row=tr, column=1).alignment = LEFT
        s.cell(row=tr, column=2, value=f"=SUM(B{start}:B{start + len(summary_plans) - 1})")
        style_total(s.cell(row=tr, column=2), "#,##0")
        s.cell(row=tr, column=3, value="100.0%")
        style_total(s.cell(row=tr, column=3), "0.0%")
        for col in range(4, len(comp_headers) + 1):
            style_total(s.cell(row=tr, column=col))

    widths = [28, 18]
    if has_py_data:
        widths.extend([18, 18, 14])
    widths.append(38)
    # Pad to at least 6 columns
    while len(widths) < 6:
        widths.append(14)
    set_widths(s, widths)


# ═════════════════════════════════════════════════════════════════════════════
# SHEET 7: DATA QUALITY
# ═════════════════════════════════════════════════════════════════════════════

def build_data_quality(wb, data):
    s = wb.create_sheet("7. Data Quality")
    plans = data.get("plans", [])

    s.merge_cells("A1:F1")
    s["A1"] = "DATA QUALITY & EXTRACTION METADATA"
    style_title(s["A1"])
    s.row_dimensions[1].height = 28

    # Only show field completeness if we have plans
    if plans:
        s["A3"] = "Field Completeness by Plan"
        style_h2(s["A3"])
        s.merge_cells(f"A3:{get_column_letter(min(len(plans) + 1, 6))}3")

        # Dynamic headers based on actual plan count (max 5 for display)
        display_plans = plans[:5]
        plan_names = [p.get("plan_name", f"Plan {i+1}")[:18] for i, p in enumerate(display_plans)]
        headers = ["Field Category"] + plan_names

        for i, h in enumerate(headers, 1):
            style_h1(s.cell(row=4, column=i, value=h))
        s.row_dimensions[4].height = 25

        all_field_checks = [
            ("Plan Identity (Name, Type)", lambda p: bool(p.get("plan_name") and p.get("plan_type"))),
            ("Roll-forward (Opening → Closing)", lambda p: has_rollforward_data(p)),
            ("Weighted Avg Exercise Price", lambda p: p.get("weighted_avg_exercise_price") is not None),
            ("Exercise Price Range", lambda p: p.get("exercise_price_range_low") is not None),
            ("Wtd Avg Contractual Life", lambda p: p.get("weighted_avg_remaining_contractual_life_years") is not None),
            ("Vesting Period", lambda p: p.get("vesting_period_years") is not None),
            ("Valuation Model", lambda p: p.get("valuation_model") is not None),
            ("Volatility / Risk-Free Rate", lambda p: bool(p.get("valuation_inputs") and (
                p["valuation_inputs"].get("volatility_pct") is not None or
                p["valuation_inputs"].get("risk_free_rate_pct") is not None
            ))),
            ("Dividend Yield", lambda p: bool(p.get("valuation_inputs") and p["valuation_inputs"].get("dividend_yield_pct") is not None)),
            ("Fair Value per Option", lambda p: bool(p.get("valuation_inputs") and p["valuation_inputs"].get("fair_value_per_option") is not None)),
            ("Tranche-Level Detail", lambda p: bool(p.get("tranches"))),
            ("Prior Year Comparatives", lambda p: bool(p.get("prior_year") and any(p["prior_year"].get(k) is not None for k in ["opening_balance", "closing_balance"]))),
            ("Performance Conditions", lambda p: has_value(p.get("performance_conditions"))),
            ("Contingent Awards", lambda p: p.get("total_contingent_awards") is not None),
        ]

        # Filter: only show fields where AT LEAST ONE plan has the data
        field_checks = [(name, fn) for name, fn in all_field_checks if any(fn(p) for p in display_plans)]

        # Also include fields where NONE have data but it's clearly relevant (commented out — keeping it clean)

        for i, (field, check_fn) in enumerate(field_checks):
            r = 5 + i
            s.cell(row=r, column=1, value=field)
            style_label(s.cell(row=r, column=1))

            for j, plan in enumerate(display_plans):
                col = 2 + j
                has_data = check_fn(plan)
                mark = "✓" if has_data else "✗"
                c = s.cell(row=r, column=col, value=mark)
                style_data(c)
                c.alignment = CENTER
                if has_data:
                    c.fill = GOOD_FILL
                    c.font = Font(name=FONT_NAME, size=10, bold=True, color="006100")
                else:
                    c.fill = BAD_FILL
                    c.font = Font(name=FONT_NAME, size=10, bold=True, color="9C0006")

        last_field_row = 5 + len(field_checks)
    else:
        last_field_row = 4

    # Validation Summary — only show items with non-zero values
    summary = data.get("_validation_summary") or {}
    all_val_items = [
        ("Total Plans Detected", summary.get("total_plans", len(plans)), "All plans on source pages captured"),
        ("Plans with Complete Roll-forward", summary.get("plans_with_complete_rollforward", 0), "Both opening and closing balance"),
        ("Plans with Valuation Inputs", summary.get("plans_with_valuation_inputs", 0), "Black-Scholes or similar"),
        ("Plans with Tranches", summary.get("plans_with_tranches", 0), "Per-grant detail extracted"),
        ("Plans with Prior Year Data", summary.get("plans_with_prior_year", 0), "Comparative columns extracted"),
        ("Plans with Exercise Price", summary.get("plans_with_exercise_price", 0), "Non-nil cost plans"),
        ("Cash-Settled Plans", summary.get("plans_cash_settled", 0), "Settled in cash, not shares"),
        ("Nil-cost Plans", summary.get("plans_nil_cost", 0), "RSU/PSP style awards"),
    ]

    # Only include items with value > 0 (and always include "Total Plans Detected")
    val_items = [(item, val, notes) for item, val, notes in all_val_items if val > 0 or item == "Total Plans Detected"]

    if val_items:
        row = last_field_row + 2
        s.cell(row=row, column=1, value="Validation Summary")
        style_h2(s.cell(row=row, column=1))
        s.merge_cells(f"A{row}:F{row}")

        row += 1
        for i, h in enumerate(["Validation Item", "Value", "Notes"], 1):
            style_h1(s.cell(row=row, column=i, value=h))

        row += 1
        for i, (item, val, notes) in enumerate(val_items):
            r = row + i
            s.cell(row=r, column=1, value=item)
            style_label(s.cell(row=r, column=1))
            s.cell(row=r, column=2, value=val)
            style_data(s.cell(row=r, column=2), "0")
            s.cell(row=r, column=2).alignment = CENTER
            s.merge_cells(f"C{r}:F{r}")
            s.cell(row=r, column=3, value=notes)
            style_data(s.cell(row=r, column=3))
            s.cell(row=r, column=3).alignment = LEFT
            s.cell(row=r, column=3).font = ITALIC_FONT

        last_row = row + len(val_items)
    else:
        last_row = last_field_row

    # Warnings (only if present)
    warnings_list = summary.get("warnings") or []
    if warnings_list:
        row = last_row + 2
        s.cell(row=row, column=1, value="Warnings & Items to Verify")
        style_h2(s.cell(row=row, column=1))
        s.merge_cells(f"A{row}:F{row}")

        for w in warnings_list:
            row += 1
            s[f"A{row}"] = f"⚠ {w}"
            s[f"A{row}"].font = LABEL_FONT
            s.merge_cells(f"A{row}:F{row}")
            s[f"A{row}"].fill = WARN_FILL

        last_row = row

    # Extraction metadata (only show fields with values)
    if "_meta" in data:
        meta = data["_meta"]

        all_meta_items = [
            ("Source PDF", meta.get("source_pdf")),
            ("Total PDF Pages", meta.get("total_pdf_pages")),
            ("Pages Processed", ", ".join(map(str, meta.get("pages_processed", []))) if meta.get("pages_processed") else None),
        ]

        # Filter: only items with values
        meta_items = [(k, v) for k, v in all_meta_items if has_value(v)]

        if meta_items:
            row = last_row + 2
            s.cell(row=row, column=1, value="Extraction Metadata")
            style_h2(s.cell(row=row, column=1))
            s.merge_cells(f"A{row}:F{row}")

            for k, v in meta_items:
                row += 1
                s[f"A{row}"] = k
                s[f"A{row}"].font = LABEL_FONT
                s.merge_cells(f"B{row}:F{row}")
                s[f"B{row}"] = v
                s[f"B{row}"].font = DATA_FONT

    set_widths(s, [32, 14, 14, 14, 14, 14])


# ═════════════════════════════════════════════════════════════════════════════
# MAIN BUILDER
# ═════════════════════════════════════════════════════════════════════════════

def build_workbook(json_path: str, output_path: str = None) -> str:
    """
    Build Excel workbook from JSON extraction output.

    Only shows fields/columns/rows with actual data — no null placeholders.

    Args:
        json_path: Path to input JSON file
        output_path: Optional output path

    Returns:
        Path to created Excel file
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if not output_path:
        company = data.get("company_name") or Path(json_path).stem
        safe_name = "".join(c for c in company if c.isalnum() or c in (" ", "_", "-")).strip()
        output_path = f"{safe_name}_options.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    print(f"Building Excel report from: {json_path}")
    plans = data.get("plans", [])
    print(f"  Plans found: {len(plans)}")

    build_executive_summary(wb, data)
    print("  ✓ Sheet 1: Executive Summary")

    build_rollforward(wb, data)
    print("  ✓ Sheet 2: Plan Roll-Forward")

    build_tranches(wb, data)
    print("  ✓ Sheet 3: Tranche Details")

    build_valuation_inputs(wb, data)
    print("  ✓ Sheet 4: Valuation Inputs")

    build_plan_descriptions(wb, data)
    print("  ✓ Sheet 5: Plan Descriptions")

    build_kpis(wb, data)
    print("  ✓ Sheet 6: KPIs & Ratios")

    build_data_quality(wb, data)
    print("  ✓ Sheet 7: Data Quality")

    wb.active = 0
    wb.save(output_path)
    print(f"\n💾 Saved: {output_path}")

    return output_path


# ═════════════════════════════════════════════════════════════════════════════
# CLI
# ═════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Convert share-based compensation JSON to Excel (clean, no null fields)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python json_to_excel.py extraction.json
  python json_to_excel.py extraction.json -o report.xlsx
        """
    )
    parser.add_argument("input", help="Path to JSON extraction output")
    parser.add_argument("-o", "--output", help="Output Excel file (default: <company>_options.xlsx)")

    args = parser.parse_args()

    if not Path(args.input).exists():
        sys.exit(f"ERROR: Input file not found: {args.input}")

    try:
        build_workbook(args.input, args.output)
    except Exception as e:
        import traceback
        traceback.print_exc()
        sys.exit(f"ERROR: {e}")


if __name__ == "__main__":
    main()